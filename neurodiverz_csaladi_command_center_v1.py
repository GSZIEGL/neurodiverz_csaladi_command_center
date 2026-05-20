from __future__ import annotations

import io
from datetime import date, datetime, time, time
from typing import Dict, List, Optional

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from supabase import create_client, Client

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    import matplotlib.pyplot as plt
except Exception:
    SimpleDocTemplate = None


def _clean_predictability_safe(value):
    text = str(value or "")
    if "| Idő:" in text:
        return text.split("| Idő:", 1)[0].strip()
    return text


def _parse_event_time_safe_from_text(value):
    text = str(value or "")
    if "| Idő:" not in text:
        return "", ""
    try:
        rng = text.split("| Idő:", 1)[1].split("|", 1)[0].strip()
        if "-" in rng:
            a, b = rng.split("-", 1)
            return a.strip(), b.strip()
    except Exception:
        pass
    return "", ""


st.set_page_config(page_title="Neurodiverz Családi Command Center v4.5.2.1", page_icon="🧩", layout="wide")

st.markdown("""
<style>
.hero{border-radius:24px;padding:24px;margin-bottom:18px;background:linear-gradient(135deg,#0f172a,#1e293b);border:1px solid rgba(255,255,255,.12)}
.hero-title{font-size:2rem;font-weight:900}.hero-sub{color:#cbd5e1;font-size:1rem;line-height:1.45}
.card{border-radius:18px;padding:18px;margin-bottom:14px;background:rgba(31,41,55,.75);border:1px solid rgba(255,255,255,.10)}
.pill{display:inline-block;padding:5px 10px;border-radius:999px;font-weight:800;margin-bottom:8px}.red{background:#7f1d1d;color:#fecaca}.yellow{background:#713f12;color:#fde68a}.green{background:#14532d;color:#bbf7d0}.blue{background:#1e3a8a;color:#bfdbfe}
</style>
""", unsafe_allow_html=True)

DAYS = ["Hétfő","Kedd","Szerda","Csütörtök","Péntek","Szombat","Vasárnap"]
WORKDAYS = ["Hétfő","Kedd","Szerda","Csütörtök","Péntek"]
PROGRAM_TYPES = {
    "Óvoda / iskola": {"base":3,"social":3,"sensory":3,"transition":2},
    "Fejlesztés / terápia": {"base":5,"social":2,"sensory":3,"transition":3},
    "Sport / mozgás": {"base":4,"social":3,"sensory":3,"transition":3},
    "Orvos / vizsgálat": {"base":7,"social":2,"sensory":4,"transition":4},
    "Bevásárlás / ügyintézés": {"base":6,"social":4,"sensory":5,"transition":4},
    "Családi program": {"base":5,"social":5,"sensory":4,"transition":3},
    "Születésnap / zsúr": {"base":8,"social":6,"sensory":6,"transition":4},
    "Utazás": {"base":5,"social":2,"sensory":3,"transition":6},
    "Otthoni nyugodt blokk": {"base":-4,"social":-2,"sensory":-3,"transition":-2},
    "Szabad játék / pihenés": {"base":-3,"social":-1,"sensory":-2,"transition":-1},
}
ENV_OPTIONS = ["Otthon","Ismerős hely","Új hely","Zsúfolt hely","Hangos hely","Sok ember","Kültér","Beltér"]
FEEDBACK_OPTIONS = {"Nyugodt":0,"Kissé fáradt":1,"Ingerlékeny":2,"Túlterhelt":3,"Meltdown közeli":4,"Meltdown volt":5}
MORNING_OPTIONS = {"Jó / kipihent":0,"Kissé nehéz indulás":1,"Fáradt":2,"Nyűgös / feszült":3,"Nagyon nehéz reggel":4}
MEAL_OPTIONS = {"Rendben evett":0,"Kicsit kevesebbet":1,"Válogatós / keveset":2,"Kimondottan problémás":3}
WEATHER_OPTIONS = ["Nincs külön hatás","Front / időjárás érzékenység","Nagy meleg","Hideg / eső","Erős szél","Nem tudjuk"]
CHALLENGE_OPTIONS = ["Nem volt","Ordítás / sírás","Verekedés / agresszió","Elvonulás / shutdown","Erős ellenállás","Iskolai/óvodai nehézség","Testvérkonfliktus","Alvás előtti kiborulás"]
CHALLENGE_TIME_OPTIONS = ["Nem volt","Reggel","Délelőtt","Dél körül","Délután","Este","Lefekvés előtt","Éjszaka"]
CHALLENGE_PHASE_OPTIONS = ["Nem volt","Program előtt","Program közben","Program után","Átálláskor / induláskor","Hazaérkezés után","Várakozás közben","Szabad játék közben","Étkezés körül","Alvás / lefekvés körül"]
CHALLENGE_LOCATION_OPTIONS = ["Nem volt","Otthon","Udvar / játszótér","Óvoda / iskola","Fejlesztés / terápia","Sport","Bolt / ügyintézés","Utazás közben","Családi programon","Ismeretlen / nem egyértelmű"]
CHALLENGE_TRIGGER_OPTIONS = ["Nem volt","Fáradtság","Éhség / szomjúság","Zaj / sok inger","Sok ember","Váratlan változás","Átállás / indulás","Várakozás","Túl hosszú program","Túl sok elvárás","Testvér / társas konfliktus","Veszteség / kudarc","Nem sikerült megnyugodni","Képernyő lezárása","Nem tudjuk"]
CHALLENGE_DURATION_OPTIONS = ["Nem volt","0–5 perc","5–15 perc","15–30 perc","30–60 perc","60+ perc","Hullámzó / visszatérő"]
SETTLE_OPTIONS = ["Nem kellett","Ölelés","Csendes szoba","Takaró alá bújás","Kedvenc zene","Mese / képernyő","Kütyüidő","Közös kuckózás","Mozgás","Tudatos légzés","Kedvenc tárgy"]

REWARD_OPTIONS = [
    "1 kocka csoki",
    "2 szem gumicukor",
    "Kakaó",
    "Kedvenc nasi",
    "Bubifújózás",
    "Játszótér",
    "Közös mese",
    "Társasjáték",
    "Közös kuckózás",
    "Séta",
    "Ő választ esti mesét",
    "Ő választ zenét",
    "Ő választ játékot"
]

HEALTH_OPTIONS = ["Nincs","Allergia","Megfázás / influenza","Hasfájás","Fejfájás","Gyógyszerváltás","ADHD tünetek erősebbek","Egyéb"]

@st.cache_resource
def get_sb() -> Client:
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_ANON_KEY"])

def restore_session(sb: Client):
    if st.session_state.get("access_token") and st.session_state.get("refresh_token"):
        try:
            sb.auth.set_session(st.session_state["access_token"], st.session_state["refresh_token"])
        except Exception:
            pass

def login_required(sb: Client):
    restore_session(sb)
    if st.session_state.get("user"):
        return st.session_state["user"]
    st.markdown('<div class="hero"><div class="hero-title">🧩 Neurodiverz Családi Command Center</div><div class="hero-sub">Felhőalapú, többfelhasználós családi stabilitástervező. Lépj be vagy regisztrálj.</div></div>', unsafe_allow_html=True)
    tab1, tab2 = st.tabs(["Belépés", "Regisztráció"])
    with tab1:
        email = st.text_input("Email", key="login_email")
        password = st.text_input("Jelszó", type="password", key="login_pw")
        if st.button("Belépés", use_container_width=True):
            try:
                res = sb.auth.sign_in_with_password({"email": email, "password": password})
                st.session_state["user"] = {"id": res.user.id, "email": res.user.email}
                st.session_state["access_token"] = res.session.access_token
                st.session_state["refresh_token"] = res.session.refresh_token
                st.rerun()
            except Exception as exc:
                st.error(f"Belépés sikertelen: {exc}")
    with tab2:
        email = st.text_input("Email", key="signup_email")
        password = st.text_input("Jelszó", type="password", key="signup_pw")
        if st.button("Regisztráció", use_container_width=True):
            try:
                sb.auth.sign_up({"email": email, "password": password})
                st.success("Regisztráció kész. Ha be van kapcsolva email megerősítés, nézd meg a postafiókod.")
            except Exception as exc:
                st.error(f"Regisztráció sikertelen: {exc}")
    st.stop()

def logout():
    for k in ["user","access_token","refresh_token"]:
        st.session_state.pop(k, None)
    st.rerun()

def q(table: str): return get_sb().table(table)

def df_from(res): return pd.DataFrame(res.data or [])

def load_children(sb): return df_from(sb.table("children").select("*").order("created_at").execute())
def load_events(sb, child_id, week): return df_from(sb.table("weekly_events").select("*").eq("child_id", child_id).eq("week_label", week).execute())
def load_all_events(sb, child_id): return df_from(sb.table("weekly_events").select("*").eq("child_id", child_id).execute())
def load_checkins(sb, child_id, week): return df_from(sb.table("daily_checkins").select("*").eq("child_id", child_id).eq("week_label", week).execute())
def load_all_checkins(sb, child_id): return df_from(sb.table("daily_checkins").select("*").eq("child_id", child_id).execute())

def safe_day_order_df(df: pd.DataFrame, day_col: str = "Nap") -> pd.DataFrame:
    if df is None or df.empty or day_col not in df.columns:
        return df
    out = df.copy()
    out["_nap_sorrend"] = out[day_col].map({d:i for i,d in enumerate(DAYS)}).fillna(99)
    return out.sort_values("_nap_sorrend").drop(columns=["_nap_sorrend"])

def collapse_checkins_by_day(checkins_df: pd.DataFrame) -> pd.DataFrame:
    # Egy napra csak egy napi rekordot hagyunk az elemzéshez.
    if checkins_df is None or checkins_df.empty or "day" not in checkins_df.columns:
        return pd.DataFrame() if checkins_df is None else checkins_df
    df = checkins_df.copy()
    if "created_at" in df.columns:
        df = df.sort_values("created_at")
    numeric_cols = ["reggeli_allapot_pont","esti_allapot_pont","napkozbeni_faradsag","alvas_minosege","etkezes_pont","kutyuidoperc","elvaras_ido","sajat_regeneracio_ido"]
    agg = {}
    for c in df.columns:
        if c == "day":
            continue
        agg[c] = "mean" if c in numeric_cols else "last"
    return df.groupby("day", as_index=False).agg(agg)

def delete_checkins_for_day(sb, child_id, week_label, day):
    return sb.table("daily_checkins").delete().eq("child_id", child_id).eq("week_label", week_label).eq("day", day).execute()

def save_single_checkin_for_day(sb, payload: Dict):
    # Új szabály: egy naphoz egy aktív check-in. Mentéskor töröljük az aznapi régieket, majd beszúrjuk az újat.
    delete_checkins_for_day(sb, payload["child_id"], payload["week_label"], payload["day"])
    return sb.table("daily_checkins").insert(payload).execute()

def cleanup_duplicate_checkins_current_week(sb, child_id, week_label) -> int:
    df = load_checkins(sb, child_id, week_label)
    if df is None or df.empty or "day" not in df.columns or "id" not in df.columns:
        return 0
    if "created_at" in df.columns:
        df = df.sort_values("created_at", ascending=False)
    deleted = 0
    for day, grp in df.groupby("day"):
        if len(grp) <= 1:
            continue
        # legfrissebb marad, többi törlődik
        for _, row in grp.iloc[1:].iterrows():
            sb.table("daily_checkins").delete().eq("id", row["id"]).execute()
            deleted += 1
    return deleted


def split_text(x): return [p.strip() for p in str(x or "").split(",") if p.strip()]
def weight(level): return 0.6 + (int(level)-1)*0.25
def clamp(x): return int(max(0, min(100, round(float(x)))))
def score_color(score, inverse=False):
    if inverse: return "🔴" if score >=75 else ("🟠" if score >=50 else "🟢")
    return "🟢" if score >=75 else ("🟠" if score >=50 else "🔴")

def calculate_event_load(event: Dict, profile: Dict) -> float:
    p = PROGRAM_TYPES[event["program_tipus"]]
    load = p["base"] + p["sensory"]*weight(profile.get("sok_ember_zaj",3)) + p["social"]*weight(profile.get("tarsas_helyzet",3)) + p["transition"]*weight(profile.get("atallas",3))
    load += max(0, float(event.get("idotartam_ora",1))-1)*1.2
    travel = int(event.get("utazas_perc",0))
    if travel >= 45: load += 5*weight(profile.get("utazas",3))
    elif travel >= 20: load += 2.5*weight(profile.get("utazas",3))
    envs = split_text(event.get("kornyezeti_tenyezok",""))
    if "Sok ember" in envs or "Zsúfolt hely" in envs: load += 5*weight(profile.get("sok_ember_zaj",3))
    if "Hangos hely" in envs: load += 4*weight(profile.get("sok_ember_zaj",3))
    if "Új hely" in envs: load += 4*weight(profile.get("rutinvalidas",3))
    if event.get("kiszamithatosag") == "Váratlan / bizonytalan": load += 6*weight(profile.get("rutinvalidas",3))
    elif event.get("kiszamithatosag") == "Részben kiszámítható": load += 3*weight(profile.get("rutinvalidas",3))
    if event["program_tipus"] in ["Otthoni nyugodt blokk","Szabad játék / pihenés"]: load -= 4*weight(profile.get("recovery_igeny",4))
    return round(load,1)

def norm_events(df):
    if df.empty: return df
    return df.rename(columns={"day":"Nap","program_tipus":"program_típus","idotartam_ora":"időtartam_óra","utazas_perc":"utazás_perc","atallas_szam":"átállás_szám","kiszamithatosag":"kiszámíthatóság","kornyezeti_tenyezok":"környezeti_tényezők","recovery_ora":"recovery_óra","szuloi_terheles":"szülői_terhelés","terhelesi_pont":"terhelési_pont"})

def norm_checkins(df):
    if df.empty: return df
    return df.rename(columns={"day":"Nap","reggeli_allapot":"Reggeli állapot","reggeli_allapot_pont":"Reggeli állapot pont","esti_allapot":"Esti állapot","esti_allapot_pont":"Esti állapot pont","napkozbeni_faradsag":"Napközbeni fáradtság","alvas_minosege":"Alvás minősége","etkezes":"Étkezés","etkezes_pont":"Étkezés pont","kutyuidoperc":"Kütyüidő perc","kulso_tenyezo":"Külső tényező","elvaras_ido":"Elvárásokkal töltött idő","sajat_regeneracio_ido":"Saját regenerációs idő","kihivas_helyzet":"Kihívást jelentő helyzet","megnyugvast_segitette":"Megnyugvást segítette","mi_segitett_szabad_szoveg":"Mi segített – saját","idoszakos_eu_allapot":"Időszakos eü. állapot"})

def normalize_events_df(events_df: pd.DataFrame) -> pd.DataFrame:
    """Supabase eseményadatok magyar, biztonságos megjelenítése.
    Kezeli a régi, időpont nélküli sorokat is.
    """
    if events_df is None or events_df.empty:
        return pd.DataFrame()

    rename = {
        "program_tipus": "program_típus",
        "idotartam_ora": "időtartam_óra",
        "utazas_perc": "utazás_perc",
        "atallas_szam": "átállás_szám",
        "kiszamithatosag": "kiszámíthatóság",
        "kornyezeti_tenyezok": "környezeti_tényezők",
        "recovery_ora": "recovery_óra",
        "szuloi_terheles": "szülői_terhelés",
        "terhelesi_pont": "terhelési_pont",
        "day": "Nap",
    }

    out = events_df.rename(columns=rename).copy()

    def _clean_local(value):
        text = str(value or "")
        if "| Idő:" in text:
            return text.split("| Idő:", 1)[0].strip()
        return text

    def _parse_time_local(value):
        text = str(value or "")
        if "| Idő:" not in text:
            return "nincs megadva"
        try:
            rng = text.split("| Idő:", 1)[1].split("|", 1)[0].strip()
            return rng.replace("-", "–")
        except Exception:
            return "nincs megadva"

    # Időpont oszlop a naptárhoz / táblához
    if "kiszámíthatóság" in out.columns:
        out["Idő"] = out["kiszámíthatóság"].apply(_parse_time_local)
        out["kiszámíthatóság"] = out["kiszámíthatóság"].apply(_clean_local)
    else:
        out["Idő"] = "nincs megadva"

    # Nap szerinti rendezés, ha van Nap oszlop
    if "Nap" in out.columns:
        out["_nap_sorrend"] = out["Nap"].map({d: i for i, d in enumerate(DAYS)})
        out = out.sort_values(["_nap_sorrend"]).drop(columns=["_nap_sorrend"])

    # A legfontosabb oszlopokat előre tesszük, a többi maradhat mögötte
    preferred = [
        "Nap", "Idő", "program_típus", "időtartam_óra", "utazás_perc",
        "átállás_szám", "kiszámíthatóság", "környezeti_tényezők",
        "recovery_óra", "szülői_terhelés", "terhelési_pont"
    ]
    ordered_cols = [c for c in preferred if c in out.columns] + [c for c in out.columns if c not in preferred]
    return out[ordered_cols]




def calculate_reward_points(checkins_df: pd.DataFrame) -> int:
    if checkins_df is None or checkins_df.empty:
        return 0
    if "day" not in checkins_df.columns:
        return 0
    return len(checkins_df["day"].dropna().unique())


def reward_level(points: int) -> str:
    if points >= 7:
        return "Arany hét"
    if points >= 5:
        return "Szuper hét"
    if points >= 3:
        return "Jó úton"
    return "Kezdő gyűjtés"


def build_day_summary(events_df, checkins_df):
    checkins_df = collapse_checkins_by_day(checkins_df)
    if events_df.empty: return pd.DataFrame()
    e = norm_events(events_df)
    s = e.groupby("Nap", as_index=False).agg(Programok=("program_típus","count"), Terhelés=("terhelési_pont","sum"), Átállások=("átállás_szám","sum"), Recovery_órák=("recovery_óra","sum"), Szülői_terhelés=("szülői_terhelés","sum"))
    s["Kockázat"] = s["Terhelés"] + s["Átállások"]*2 + s["Szülői_terhelés"] - s["Recovery_órák"]*4
    s["_o"] = s["Nap"].map({d:i for i,d in enumerate(DAYS)})
    s = s.sort_values("_o").drop(columns="_o")
    if not checkins_df.empty:
        c = norm_checkins(checkins_df)
        cols = [x for x in ["Nap","Esti állapot pont","Esti állapot","Reggeli állapot pont","Étkezés pont","Kütyüidő perc","Napközbeni fáradtság","Elvárásokkal töltött idő","Saját regenerációs idő"] if x in c.columns]
        s = s.merge(c[cols], on="Nap", how="left")
        for col,m in [("Reggeli állapot pont",3),("Esti állapot pont",4),("Étkezés pont",3),("Napközbeni fáradtság",2.5),("Elvárásokkal töltött idő",1.2)]:
            if col in s: s["Kockázat"] += s[col].fillna(0)*m
        if "Kütyüidő perc" in s: s["Kockázat"] += (s["Kütyüidő perc"].fillna(0)/30)*2
        if "Saját regenerációs idő" in s: s["Kockázat"] -= s["Saját regenerációs idő"].fillna(0)*2
    return s

def generate_insights(day_summary, profile, events_df, checkins_df):
    out=[]
    if day_summary.empty: return [{"Szint":"INFO","Téma":"Nincs még adat","Mit látunk?":"Adj hozzá heti programokat.","Mit érdemes tenni?":"Kezdd a fix programokkal: óvoda/iskola, fejlesztés, sport, családi programok."}]
    overload = day_summary[day_summary["Kockázat"]>=50]
    if len(overload): out.append({"Szint":"FIGYELMEZTETÉS","Téma":"Túlterhelt napok","Mit látunk?":f"Magasabb idegrendszeri terhelés látszik: {', '.join(overload['Nap'])}.","Mit érdemes tenni?":"Kevesebb plusz program, több előrejelzés és több nyugodt blokk javasolt."})
    low_rec = day_summary[day_summary["Recovery_órák"]<1]
    if len(low_rec)>=3: out.append({"Szint":"FIGYELMEZTETÉS","Téma":"Kevés recovery","Mit látunk?":"Több napon kevés valódi lecsendesedési idő látszik.","Mit érdemes tenni?":"Tegyél be legalább 2–3 rövid, védett nyugalmi blokkot."})
    ordered = safe_day_order_df(day_summary.groupby("Nap", as_index=False).agg(lambda x: x.mean() if pd.api.types.is_numeric_dtype(x) else x.iloc[-1])).fillna(0)
    streak=max_streak=0
    for high in (ordered["Kockázat"]>=45):
        streak = streak+1 if high else 0; max_streak=max(max_streak,streak)
    if max_streak>=3: out.append({"Szint":"KRITIKUS","Téma":"Egymást követő nehéz napok","Mit látunk?":"Legalább 3 egymást követő magasabb terhelésű nap látszik.","Mit érdemes tenni?":"Tegyél recovery blokkot a sorozat közepére vagy végére."})
    if int(profile.get("screen_sensitivity",3))>=4 and not checkins_df.empty and "kutyuidoperc" in checkins_df:
        if len(checkins_df[checkins_df["kutyuidoperc"]>=45]): out.append({"Szint":"FIGYELMEZTETÉS","Téma":"Képernyőidő figyelendő","Mit látunk?":"A profil szerint a képernyő érzékeny pont, és volt magasabb kütyüidő.","Mit érdemes tenni?":"Teszteld, hogy a képernyőidő csökkentése javítja-e az esti állapotot vagy a másnapot."})
    if not out: out.append({"Szint":"INFO","Téma":"Kezelhető hét","Mit látunk?":"Nincs kiugró túlterhelési jel.","Mit érdemes tenni?":"Tartsd meg a nyugodt blokkokat, és rögzíts check-int."})
    return out

def engagement(checkins):
    n = checkins["day"].nunique() if not checkins.empty and "day" in checkins else 0
    if n>=7: return "🏆 Hibátlan hét: minden naphoz van check-in."
    if n>=5: return "🌟 Nagyon jó: már legalább 5 napot rögzítettetek."
    if n>=3: return "✅ Szép munka: már látszanak a mintázatok."
    if n>=1: return "👍 Jó kezdés: már egy check-in is hasznos."
    return "🧩 Kezdd egyetlen napi check-innel. Nem kell tökéletesen vezetni."


def build_deeper_conclusions(summary, profile, events, checkins):
    conclusions=[]
    if summary.empty: return conclusions
    hardest=summary.sort_values("Kockázat",ascending=False).iloc[0]
    conclusions.append({"Cím":"A hét legnehezebb pontja","Következtetés":f"A legnagyobb terhelés most ezen a napon látszik: {hardest['Nap']}.","Javaslat":"Ha csak egy dolgon változtattok, először ezt a napot érdemes könnyíteni vagy több recoveryt tenni utána."})
    if "Kütyüidő perc" in summary.columns and summary["Kütyüidő perc"].notna().sum()>=2 and int(profile.get("screen_sensitivity",3))>=3:
        high=summary[summary["Kütyüidő perc"].fillna(0)>=45]
        if len(high): conclusions.append({"Cím":"Képernyőidő mintázat","Következtetés":"Volt olyan nap, amikor magasabb képernyőidő jelent meg, és a profil alapján ez érzékeny terület lehet.","Javaslat":"Érdemes 1–2 hétig figyelni, hogy a plusz képernyőidő után romlik-e az esti állapot vagy a másnap."})
    if "Étkezés pont" in summary.columns and summary["Étkezés pont"].fillna(0).max()>=2:
        conclusions.append({"Cím":"Étkezési jelzés","Következtetés":"Az étkezésben látszik eltérés valamelyik napon. Ez sok neurodivergens gyereknél korai fáradási jel lehet.","Javaslat":"Érdemes figyelni, hogy az étkezési változás együtt jár-e fáradtsággal, ingerlékenységgel vagy több képernyőigénnyel."})
    if summary["Recovery_órák"].sum()<4:
        conclusions.append({"Cím":"Kevés visszatöltő idő","Következtetés":"A héten kevés tervezett nyugodt blokk látszik.","Javaslat":"Tegyél be legalább két rövidebb, védett lecsendesedési időt. A cél nem a programok törlése, hanem a hét fenntarthatóbbá tétele."})
    if not checkins.empty and "megnyugvast_segitette" in checkins.columns:
        helped=[]
        for v in checkins["megnyugvast_segitette"].dropna().astype(str): helped += [x.strip() for x in v.split(",") if x.strip()]
        if helped:
            top=pd.Series(helped).value_counts().index[0]
            conclusions.append({"Cím":"Ami működni látszik","Következtetés":f"A visszajelzések alapján ez többször megjelent segítő stratégiaként: {top}.","Javaslat":"Ezt érdemes tudatosan előre betervezni a nehezebb napok elé vagy után."})
    return conclusions[:6]

def pdf_safe(x):
    return str(x or "").replace("ő","ö").replace("Ő","Ö").replace("ű","ü").replace("Ű","Ü")


def classify_week_type(day_summary: pd.DataFrame) -> Dict[str, str]:
    """Heti összkép egyszerű, szülőbarát besorolással."""
    if day_summary.empty:
        return {"Típus": "Nincs elég adat", "Magyarázat": "Még nincs elég heti adat a besoroláshoz."}

    avg_risk = day_summary["Kockázat"].mean()
    max_risk = day_summary["Kockázat"].max()
    std_risk = day_summary["Kockázat"].std() if len(day_summary) > 1 else 0
    recovery_sum = day_summary.get("Recovery_órák", pd.Series(dtype=float)).sum()

    ordered = safe_day_order_df(day_summary.groupby("Nap", as_index=False).agg(lambda x: x.mean() if pd.api.types.is_numeric_dtype(x) else x.iloc[-1])).fillna(0)
    ordered["Kockázat"] = ordered["Kockázat"].fillna(0)
    second_half = ordered[ordered["Nap"].isin(["Csütörtök", "Péntek", "Szombat", "Vasárnap"])]["Kockázat"].mean()
    first_half = ordered[ordered["Nap"].isin(["Hétfő", "Kedd", "Szerda"])]["Kockázat"].mean()

    if max_risk >= 75 or avg_risk >= 55:
        return {
            "Típus": "Túlterhelt hét",
            "Magyarázat": "A hét egészében vagy legalább egy napon magas idegrendszeri terhelés látszik."
        }
    if second_half > first_half + 15:
        return {
            "Típus": "Fáradó hét",
            "Magyarázat": "A hét második felére emelkedik a terhelés, ami fokozatos kifáradásra utalhat."
        }
    if std_risk >= 20:
        return {
            "Típus": "Hullámzó hét",
            "Magyarázat": "A napok terhelése nagyon eltérő. Ilyenkor nehéz lehet kiszámítható ritmust tartani."
        }
    if recovery_sum < 4:
        return {
            "Típus": "Recovery-hiányos hét",
            "Magyarázat": "Kevés tervezett visszatöltő idő látszik a héten."
        }
    return {
        "Típus": "Alapvetően stabil hét",
        "Magyarázat": "A jelenlegi adatok alapján nincs erős túlterhelési minta."
    }


def _safe_mean(series):
    try:
        vals = pd.to_numeric(series, errors="coerce").dropna()
        return vals.mean() if len(vals) else np.nan
    except Exception:
        return np.nan



def parse_challenge_details(value: object) -> Dict[str, str]:
    """A régi és új 'kihivas_helyzet' mezőből is olvasható részletek."""
    text = "" if value is None else str(value)
    result = {
        "alap": text.strip(),
        "időszak": "",
        "kapcsolódás": "",
        "helyzet": "",
        "kiváltó": "",
        "időtartam": "",
    }
    if "|" not in text:
        return result
    parts = [p.strip() for p in text.split("|")]
    result["alap"] = parts[0] if parts else ""
    for p in parts[1:]:
        if ":" in p:
            k, v = p.split(":", 1)
            k = k.strip().lower()
            v = v.strip()
            if "időszak" in k:
                result["időszak"] = v
            elif "kapcsolódás" in k:
                result["kapcsolódás"] = v
            elif "helyzet" in k:
                result["helyzet"] = v
            elif "kiváltó" in k:
                result["kiváltó"] = v
            elif "időtartam" in k:
                result["időtartam"] = v
    return result


def build_prognosis_engine(day_summary: pd.DataFrame, profile: Dict, events_df: pd.DataFrame, checkins_df: pd.DataFrame) -> pd.DataFrame:
    """Egyszerű, szabályalapú előrejelző réteg a következő napokra / hétre."""
    rows = []
    if day_summary.empty:
        return pd.DataFrame([{
            "Terület": "Nincs elég adat",
            "Előrejelzés": "Adj hozzá programokat és néhány check-int, és a rendszer óvatos prognózist ad.",
            "Mire figyelj?": "A prognózis nem diagnózis, csak tervezési segítség.",
            "Megelőző lépés": "Kezdd napi 1 rövid check-innel."
        }])

    ordered = safe_day_order_df(day_summary.groupby("Nap", as_index=False).agg(lambda x: x.mean() if pd.api.types.is_numeric_dtype(x) else x.iloc[-1])).fillna(0)
    ordered["Kockázat"] = pd.to_numeric(ordered["Kockázat"], errors="coerce").fillna(0)
    ordered["Recovery_órák"] = pd.to_numeric(ordered.get("Recovery_órák", 0), errors="coerce").fillna(0)

    # Következő magas kockázatú napok
    risk_days = ordered[ordered["Kockázat"] >= 45]
    if not risk_days.empty:
        rows.append({
            "Terület": "Következő nehezebb napok",
            "Előrejelzés": f"Ezek a napok előre terheltebbnek tűnnek: {', '.join(risk_days['Nap'].astype(str).tolist())}.",
            "Mire figyelj?": "Ezeken a napokon könnyebben jöhet túlterhelés, főleg ha kevés a lecsendesedési idő.",
            "Megelőző lépés": "Előre jelezd a nap menetét, és tervezz rövid, elvárásmentes blokkot a legterheltebb program után."
        })

    # Halmozódás / sorozat
    streak = 0
    max_streak = 0
    for val in ordered["Kockázat"] >= 40:
        streak = streak + 1 if val else 0
        max_streak = max(max_streak, streak)
    if max_streak >= 2:
        rows.append({
            "Terület": "Terhelés halmozódása",
            "Előrejelzés": "Több egymást követő figyelendő nap látszik.",
            "Mire figyelj?": "Ilyenkor nem mindig az első nehéz nap borít, hanem a második-harmadik.",
            "Megelőző lépés": "A sorozat közepére tegyél egy alacsony ingerű délutánt vagy rövidebb programot."
        })

    # Hét végi kifáradás
    first = ordered[ordered["Nap"].isin(["Hétfő","Kedd","Szerda"])]["Kockázat"].mean()
    second = ordered[ordered["Nap"].isin(["Csütörtök","Péntek","Szombat","Vasárnap"])]["Kockázat"].mean()
    if pd.notna(first) and pd.notna(second) and second > first + 10:
        rows.append({
            "Terület": "Hétvégi / hétvége előtti fáradás",
            "Előrejelzés": "A hét második fele nehezebbnek tűnik.",
            "Mire figyelj?": "A csütörtök-péntek körüli fáradás sokszor késleltetve jelenik meg.",
            "Megelőző lépés": "Csütörtök után érdemes kevesebb új helyzetet és több kiszámítható rutint hagyni."
        })

    # Challenge részletek alapján trigger prognózis
    if not checkins_df.empty and "kihivas_helyzet" in checkins_df.columns:
        parsed = checkins_df["kihivas_helyzet"].dropna().apply(parse_challenge_details)
        triggers = [x.get("kiváltó","") for x in parsed if x.get("kiváltó","") and x.get("kiváltó","") != "Nem volt"]
        phases = [x.get("kapcsolódás","") for x in parsed if x.get("kapcsolódás","") and x.get("kapcsolódás","") != "Nem volt"]
        times = [x.get("időszak","") for x in parsed if x.get("időszak","") and x.get("időszak","") != "Nem volt"]
        if triggers:
            top_trigger = pd.Series(triggers).value_counts().index[0]
            rows.append({
                "Terület": "Visszatérő kiváltó ok",
                "Előrejelzés": f"A rögzített adatok alapján ez figyelendő trigger lehet: {top_trigger}.",
                "Mire figyelj?": "Nem biztos, hogy ez az ok, de érdemes a következő hetekben külön megfigyelni.",
                "Megelőző lépés": "Ha ez a trigger várható, előtte legyen rövid előrejelzés, utána pedig visszatöltő blokk."
            })
        if phases:
            top_phase = pd.Series(phases).value_counts().index[0]
            rows.append({
                "Terület": "Mikor jelenik meg a nehézség?",
                "Előrejelzés": f"A nehéz helyzetek gyakran ehhez kapcsolódhatnak: {top_phase}.",
                "Mire figyelj?": "A program előtti, közbeni és utáni nehézség más-más megelőzést igényel.",
                "Megelőző lépés": "A kapcsolódó időszakba tegyél több kiszámíthatóságot és kevesebb elvárást."
            })
        if times:
            top_time = pd.Series(times).value_counts().index[0]
            rows.append({
                "Terület": "Napszakos érzékenység",
                "Előrejelzés": f"A rögzített nehézségek gyakrabban ebben az időszakban jelentek meg: {top_time}.",
                "Mire figyelj?": "Lehet, hogy a napszak önmagában is kapacitáscsökkenést jelez.",
                "Megelőző lépés": "Ebben az időszakban érdemes kevesebb plusz döntést és átállást kérni."
            })

    # Screen prognózis
    if not checkins_df.empty and "kutyuidoperc" in checkins_df.columns and int(profile.get("screen_sensitivity", 3)) >= 4:
        avg_screen = pd.to_numeric(checkins_df["kutyuidoperc"], errors="coerce").dropna().mean()
        if pd.notna(avg_screen) and avg_screen >= 45:
            rows.append({
                "Terület": "Képernyőérzékenység",
                "Előrejelzés": "A profil és az adatok alapján a képernyőidő külön figyelendő.",
                "Mire figyelj?": "A hatás nem mindig azonnali; lehet esti vagy másnapi feszültség.",
                "Megelőző lépés": "Ne teljes tiltással kezdj, hanem tesztelj 1 héten át rövidebb vagy korábbi képernyőidőt."
            })

    if not rows:
        rows.append({
            "Terület": "Óvatos prognózis",
            "Előrejelzés": "A jelenlegi adatok alapján nincs erős előrejelző jel.",
            "Mire figyelj?": "Ez nem azt jelenti, hogy biztosan könnyű hét lesz, csak nincs még elég markáns minta.",
            "Megelőző lépés": "Folytasd a rövid check-ineket; 2–3 hét után pontosabb lesz a rendszer."
        })

    return pd.DataFrame(rows[:8])


def build_challenge_pattern_table(checkins_df: pd.DataFrame) -> pd.DataFrame:
    """Kihívást jelentő helyzetek áttekintése strukturáltan."""
    if checkins_df.empty or "kihivas_helyzet" not in checkins_df.columns:
        return pd.DataFrame()
    rows = []
    for _, r in checkins_df.iterrows():
        details = parse_challenge_details(r.get("kihivas_helyzet", ""))
        if details.get("alap", "") and details.get("alap", "") != "Nem volt":
            rows.append({
                "Nap": r.get("day", ""),
                "Helyzet": details.get("alap", ""),
                "Időszak": details.get("időszak", ""),
                "Kapcsolódás": details.get("kapcsolódás", ""),
                "Helyszín": details.get("helyzet", ""),
                "Kiváltó": details.get("kiváltó", ""),
                "Időtartam": details.get("időtartam", ""),
            })
    return pd.DataFrame(rows)


def build_neurodiverz_insight_engine(
    day_summary: pd.DataFrame,
    profile: Dict,
    events_df: pd.DataFrame,
    checkins_df: pd.DataFrame
) -> Dict[str, pd.DataFrame]:
    """Mélyebb, több rétegű insight motor AI nélkül."""
    weekly_rows = []
    day_rows = []
    pattern_rows = []
    positive_rows = []
    micro_rows = []

    if day_summary.empty:
        empty = pd.DataFrame(columns=["Típus", "Megállapítás", "Miért fontos?", "Javaslat"])
        return {
            "heti_osszkep": empty,
            "napi_fokusz": empty,
            "mintazatok": empty,
            "pozitiv_jelek": empty,
            "mikro_javaslatok": empty,
        }

    week_type = classify_week_type(day_summary)
    weekly_rows.append({
        "Típus": week_type["Típus"],
        "Megállapítás": week_type["Magyarázat"],
        "Miért fontos?": "A heti ritmus sokszor fontosabb, mint egyetlen program önmagában.",
        "Javaslat": "A cél nem minden terhelés megszüntetése, hanem a nehéz napok köré elég visszatöltő időt tenni."
    })

    ordered = safe_day_order_df(day_summary.groupby("Nap", as_index=False).agg(lambda x: x.mean() if pd.api.types.is_numeric_dtype(x) else x.iloc[-1])).fillna(0)
    ordered["Kockázat"] = pd.to_numeric(ordered["Kockázat"], errors="coerce").fillna(0)
    ordered["Recovery_órák"] = pd.to_numeric(ordered.get("Recovery_órák", 0), errors="coerce").fillna(0)

    # Több nap fókusz: top 3 nem csak legmagasabb
    top_days = ordered[ordered["Kockázat"] > 0].sort_values("Kockázat", ascending=False).head(3)
    for _, row in top_days.iterrows():
        level = "Magas" if row["Kockázat"] >= 55 else ("Közepes" if row["Kockázat"] >= 35 else "Figyelendő")
        day_rows.append({
            "Nap": row["Nap"],
            "Fókusz": f"{level} terhelés",
            "Mit látunk?": f"{row['Nap']} kockázati pontja: {row['Kockázat']:.0f}.",
            "Javaslat": "Ezen a napon érdemes előre jelezni a programokat, és legalább egy rövid lecsendesedési blokkot biztosítani."
        })

    # Hét második felének kifáradása
    first = ordered[ordered["Nap"].isin(["Hétfő", "Kedd", "Szerda"])]["Kockázat"].mean()
    second = ordered[ordered["Nap"].isin(["Csütörtök", "Péntek", "Szombat", "Vasárnap"])]["Kockázat"].mean()
    if pd.notna(first) and pd.notna(second) and second > first + 12:
        pattern_rows.append({
            "Minta": "Hét második felére emelkedő terhelés",
            "Mit látunk?": "A hét második fele nehezebbnek tűnik, mint az első.",
            "Miért fontos?": "Sok neurodivergens gyereknél a hét végére csökken a tartalék, még akkor is, ha minden nap külön-külön kezelhetőnek tűnik.",
            "Javaslat": "Csütörtök-péntek környékére érdemes könnyebb délutánt vagy több saját regenerációs időt hagyni."
        })

    # Recovery minta
    low_rec_days = ordered[(ordered["Recovery_órák"] < 1) & (ordered["Kockázat"] > 25)]
    if len(low_rec_days) >= 2:
        pattern_rows.append({
            "Minta": "Nehéz napok kevés recoveryvel",
            "Mit látunk?": f"Több terheltebb napon kevés visszatöltő idő látszik: {', '.join(low_rec_days['Nap'].astype(str).tolist())}.",
            "Miért fontos?": "A gond sokszor nem a program, hanem az, hogy utána nincs idegrendszeri lecsengés.",
            "Javaslat": "A nehezebb programok után 20–40 perc elvárásmentes blokk sokat segíthet."
        })

    # Checkin alapú minták
    if not checkins_df.empty:
        c = checkins_df.copy()
        # expected raw supabase columns
        if "kutyuidoperc" in c.columns and "esti_allapot_pont" in c.columns:
            high_screen = c[pd.to_numeric(c["kutyuidoperc"], errors="coerce").fillna(0) >= 45]
            if len(high_screen) >= 1 and int(profile.get("screen_sensitivity", 3)) >= 3:
                pattern_rows.append({
                    "Minta": "Képernyőidő figyelendő",
                    "Mit látunk?": "Volt magasabb képernyőidővel járó nap.",
                    "Miért fontos?": "Egyes gyerekeknél a plusz képernyőidő nem aznap, hanem este vagy másnap jelentkezik terhelésként.",
                    "Javaslat": "Érdemes 1–2 hétig tesztelni: kevesebb képernyő a nehezebb napokon, és figyelni az esti állapotot."
                })

        if "etkezes_pont" in c.columns:
            meal_mean = _safe_mean(c["etkezes_pont"])
            if pd.notna(meal_mean) and meal_mean >= 1.4:
                pattern_rows.append({
                    "Minta": "Étkezés mint korai jel",
                    "Mit látunk?": "Több napon nem teljesen stabil az étkezés.",
                    "Miért fontos?": "Az étkezés romlása sok családnál hamarabb jelez túlterhelést, mint a látványos kiborulás.",
                    "Javaslat": "Nézzétek meg, hogy az étkezési eltérés együtt jár-e rosszabb alvással, fáradtsággal vagy képernyőigénnyel."
                })

        if "reggeli_allapot_pont" in c.columns and "esti_allapot_pont" in c.columns:
            morning_avg = _safe_mean(c["reggeli_allapot_pont"])
            evening_avg = _safe_mean(c["esti_allapot_pont"])
            if pd.notna(morning_avg) and pd.notna(evening_avg) and evening_avg > morning_avg + 1:
                pattern_rows.append({
                    "Minta": "Nap közbeni lemerülés",
                    "Mit látunk?": "Az esti állapot érezhetően rosszabb, mint a reggeli indulás.",
                    "Miért fontos?": "Ez arra utalhat, hogy a nap közbeni elvárások, ingerek vagy átállások fokozatosan merítik a gyermeket.",
                    "Javaslat": "Érdemes napközben is keresni mini visszatöltő pontokat, nem csak estére hagyni a lecsendesedést."
                })

        if "megnyugvast_segitette" in c.columns:
            helped = []
            for value in c["megnyugvast_segitette"].dropna().astype(str):
                helped += [x.strip() for x in value.split(",") if x.strip()]
            if helped:
                top_help = pd.Series(helped).value_counts().index[0]
                positive_rows.append({
                    "Pozitív minta": "Van működő megnyugvási stratégia",
                    "Mit látunk?": f"Ez többször is segítő tényezőként jelent meg: {top_help}.",
                    "Miért jó jel?": "Nem csak a nehézséget látjuk, hanem azt is, mi segít visszaterelni a rendszert.",
                    "Javaslat": "Ezt érdemes előre betervezni a nehezebb napok elé vagy után, nem csak utólag használni."
                })

        if "sajat_regeneracio_ido" in c.columns:
            rec_avg = _safe_mean(c["sajat_regeneracio_ido"])
            if pd.notna(rec_avg) and rec_avg >= 1.5:
                positive_rows.append({
                    "Pozitív minta": "Megjelent saját regenerációs idő",
                    "Mit látunk?": "Több napon volt saját tempójú pihenés vagy szabad játék.",
                    "Miért jó jel?": "A visszatöltő idő védőfaktor lehet a hét során.",
                    "Javaslat": "Ezt érdemes tudatosan megtartani, még akkor is, ha látszólag 'nem történik semmi'."
                })

    # Program alapú minták
    if not events_df.empty:
        e = events_df.copy()
        if "atallas_szam" in e.columns:
            high_transition_events = e[pd.to_numeric(e["atallas_szam"], errors="coerce").fillna(0) >= 3]
            if len(high_transition_events) >= 2 and int(profile.get("atallas", 3)) >= 3:
                pattern_rows.append({
                    "Minta": "Átállási terhelés",
                    "Mit látunk?": "Több program körül sok átállás jelenik meg.",
                    "Miért fontos?": "Az átállás sokszor láthatatlan terhelés: indulás, öltözés, érkezés, váltás, hazaérés.",
                    "Javaslat": "A sok átállásos napokon segíthet a vizuális sorrend, fix indulási rutin és plusz idő."
                })

        if "kornyezeti_tenyezok" in e.columns:
            crowded = e[e["kornyezeti_tenyezok"].fillna("").str.contains("Sok ember|Zsúfolt hely|Hangos hely", regex=True)]
            if len(crowded) >= 2 and int(profile.get("sok_ember_zaj", 3)) >= 3:
                pattern_rows.append({
                    "Minta": "Ingergazdag környezetek halmozódása",
                    "Mit látunk?": "Több zajos, zsúfolt vagy sok emberrel járó helyzet szerepel a héten.",
                    "Miért fontos?": "Ezek önmagukban is merítőek lehetnek, de egymás után különösen leterhelők.",
                    "Javaslat": "Ha nem lehet elkerülni, előtte/utána legyen kiszámítható, alacsony ingerű blokk."
                })

    # Mikro-javaslatok, mindig legyen néhány konkrét
    if not day_rows:
        micro_rows.append({
            "Javaslat": "Tartsátok meg a jelenlegi ritmust",
            "Mikor?": "A hét egészében",
            "Hogyan?": "A meglévő nyugodt blokkok és előrejelzések megtartása most fontosabb lehet, mint új szabályok bevezetése."
        })
    else:
        for row in day_rows[:3]:
            micro_rows.append({
                "Javaslat": f"{row['Nap']} könnyítése",
                "Mikor?": row["Nap"],
                "Hogyan?": "Egy plusz program elhagyása, rövidebb ott tartózkodás, vagy program után 20–30 perc elvárásmentes idő."
            })

    if len(positive_rows) == 0:
        positive_rows.append({
            "Pozitív minta": "Már maga a rögzítés is segítség",
            "Mit látunk?": "A hét adatai alapján elkezdhető a mintázatok keresése.",
            "Miért jó jel?": "Nem kell tökéletesen vezetni: néhány adatpont is segíthet a családnak kívülről ránézni a hétre.",
            "Javaslat": "A következő héten elég napi 1 gyors check-inre törekedni."
        })

    return {
        "heti_osszkep": pd.DataFrame(weekly_rows),
        "napi_fokusz": pd.DataFrame(day_rows),
        "mintazatok": pd.DataFrame(pattern_rows),
        "pozitiv_jelek": pd.DataFrame(positive_rows),
        "mikro_javaslatok": pd.DataFrame(micro_rows),
    }


def build_visual_pdf_report(profile, events, summary, insights, checkins, week_label):
    if SimpleDocTemplate is None: return None
    output=io.BytesIO()
    doc=SimpleDocTemplate(output,pagesize=A4,rightMargin=1.2*cm,leftMargin=1.2*cm,topMargin=1.1*cm,bottomMargin=1.1*cm)
    styles=getSampleStyleSheet(); title=ParagraphStyle("t",parent=styles["Title"],fontSize=18,leading=22); h2=ParagraphStyle("h",parent=styles["Heading2"],fontSize=12,textColor=colors.HexColor("#1E3A8A")); body=ParagraphStyle("b",parent=styles["Normal"],fontSize=8.5,leading=11); small=ParagraphStyle("s",parent=styles["Normal"],fontSize=7.5,leading=9)
    def P(v,style=body): return Paragraph(pdf_safe(v), style)
    story=[P(f"Neurodiverz családi heti riport – {profile.get('nickname','Gyermek')}",title),P(f"Hét: {week_label} · Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M')}",small),Spacer(1,.25*cm)]
    if summary.empty:
        story.append(P("Nincs elég adat a heti riporthoz.")); 
    # Összes stabilitási következtetés exportálása
    story.append(P("Összes stabilitási következtetés", h2))

    try:
        insights_df = pd.DataFrame(generate_insights(day_summary, profile, events_df, checkins_df))
        if insights_df is not None and not insights_df.empty:
            for _, row in insights_df.iterrows():
                title = str(row.get("Téma", row.get("Terület", "Következtetés")))
                text = str(row.get("Következtetés", row.get("Előrejelzés", "")))
                action = str(row.get("Javaslat", row.get("Megelőző lépés", "")))

                story.append(P(f"<b>{title}</b>", small))
                if text:
                    story.append(P(text, small))
                if action:
                    story.append(P(f"Javaslat: {action}", small))
                story.append(Spacer(1, 0.08*cm))
        else:
            story.append(P("Még nincs elég adat részletes stabilitási következtetésekhez.", small))
    except Exception as export_exc:
        story.append(P(f"Export következtetés-hiba: {export_exc}", small))


    doc.build(story); return output.getvalue()
    risk=clamp(summary["Kockázat"].sum()/330*100); stability=clamp(100-risk+summary["Recovery_órák"].sum()*2); maxday=summary.sort_values("Kockázat",ascending=False).iloc[0]; checkdays=checkins["day"].nunique() if not checkins.empty and "day" in checkins else 0
    data=[[P("Heti stabilitás",small),P("Túlterhelési kockázat",small),P("Legnehezebb nap",small),P("Check-in napok",small)],[P(f"{stability}/100"),P(f"{risk}/100"),P(maxday["Nap"]),P(str(checkdays))]]
    t=Table(data,colWidths=[4.2*cm]*4); t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E3A8A")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("BACKGROUND",(0,1),(-1,1),colors.HexColor("#EFF6FF")),("GRID",(0,0),(-1,-1),.3,colors.HexColor("#CBD5E1")),("VALIGN",(0,0),(-1,-1),"TOP")]))
    story += [t,Spacer(1,.35*cm)]
    # chart as PNG
    fig,ax=plt.subplots(figsize=(7,2.6)); vals=summary["Kockázat"].fillna(0); labs=summary["Nap"]
    ax.bar(labs, vals); ax.set_title("Napi idegrendszeri terhelés / kockázat"); ax.set_ylabel("pont"); ax.tick_params(axis='x', rotation=25); fig.tight_layout()
    img=io.BytesIO(); fig.savefig(img,format='png',dpi=160); plt.close(fig); img.seek(0); story.append(Image(img,width=16*cm,height=5.8*cm)); story.append(Spacer(1,.3*cm))
    story.append(P("Fő következtetések",h2))
    for i,item in enumerate(build_deeper_conclusions(summary,profile,events,checkins)[:4],1):
        story.append(P(f"{i}. {item['Cím']}",body)); story.append(P(f"Következtetés: {item['Következtetés']}",small)); story.append(P(f"Javaslat: {item['Javaslat']}",small)); story.append(Spacer(1,.12*cm))
    story.append(P("Szülőbarát insightok",h2))
    for _,r in insights.head(4).iterrows():
        story.append(P(f"{r.get('Szint','')} – {r.get('Téma','')}",body)); story.append(P(f"Mit látunk? {r.get('Mit látunk?','')}",small)); story.append(P(f"Mit érdemes tenni? {r.get('Mit érdemes tenni?','')}",small)); story.append(Spacer(1,.1*cm))
    story.append(P("Megjegyzés",h2)); story.append(P("Ez az eszköz nem diagnosztikai vagy egészségügyi rendszer. Célja a családi terhelés, rutin, átállások, alvás/étkezés/képernyő és recovery tudatosabb tervezése.",small))
    doc.build(story); return output.getvalue()

def export_excel(profile, events, summary, insights, checkins):
    bio=io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as w:
        pd.DataFrame([profile]).to_excel(w,index=False,sheet_name="Gyermekprofil")
        norm_events(events).to_excel(w,index=False,sheet_name="Programok")
        summary.to_excel(w,index=False,sheet_name="Napi összegzés")
        insights.to_excel(w,index=False,sheet_name="Insightok")
        norm_checkins(checkins).to_excel(w,index=False,sheet_name="Check-inek")
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        for ws in w.book.worksheets:
            for cell in ws[1]: cell.fill=PatternFill("solid",fgColor="1E3A8A"); cell.font=Font(color="FFFFFF",bold=True); cell.alignment=Alignment(wrap_text=True)
            for i,col in enumerate(ws.columns,1):
                ws.column_dimensions[get_column_letter(i)].width=26
                for cell in col: cell.alignment=Alignment(wrap_text=True,vertical="top")
    return bio.getvalue()

def copy_week(sb, child_id, source_week, target_week):
    src=load_events(sb,child_id,source_week); n=0
    for _,r in src.iterrows():
        payload={"child_id":child_id,"week_label":target_week,"day":r["day"],"program_tipus":r["program_tipus"],"idotartam_ora":float(r.get("idotartam_ora",1)),"utazas_perc":int(r.get("utazas_perc",0)),"atallas_szam":int(r.get("atallas_szam",1)),"kiszamithatosag":r.get("kiszamithatosag","Kiszámítható"),"kornyezeti_tenyezok":r.get("kornyezeti_tenyezok",""),"recovery_ora":float(r.get("recovery_ora",0)),"szuloi_terheles":int(r.get("szuloi_terheles",2)),"terhelesi_pont":float(r.get("terhelesi_pont",0))}
        sb.table("weekly_events").insert(payload).execute(); n+=1
    return n

sb=get_sb(); user=login_required(sb)
with st.sidebar:
    st.write(f"Belépve: **{user['email']}**")
    if st.button("Kijelentkezés", use_container_width=True): logout()

st.markdown('<div class="hero"><div class="hero-title">🧩 Neurodiverz Családi Command Center v4.5.2.1</div><div class="hero-sub">Felhőalapú, többfelhasználós stabilitástervező. Belépés után bárhonnan elérhető, és több hét adataiból kezd mintázatokat mutatni.</div></div>', unsafe_allow_html=True)

children=load_children(sb)
with st.sidebar:
    st.header("Gyermek kiválasztása")
    if children.empty:
        selected_child_id=None; st.info("Még nincs gyermekprofil.")
    else:
        opts=dict(zip(children["nickname"],children["id"])); name=st.selectbox("Gyermek",list(opts.keys())); selected_child_id=opts[name]
    week_label=st.text_input("Hét azonosító", value=f"{date.today().isocalendar().year}-W{date.today().isocalendar().week:02d}")


def _time_to_minutes(value):
    try:
        if not value:
            return 9999
        h, m = str(value).split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return 9999


def _event_time_text(start_time_value, end_time_value):
    if start_time_value is None or end_time_value is None:
        return ""
    try:
        return f" | Idő: {start_time_value.strftime('%H:%M')}-{end_time_value.strftime('%H:%M')}"
    except Exception:
        return ""


def _parse_event_time(row):
    text = str(row.get("kiszamithatosag", row.get("kiszámíthatóság", "")) or "")
    if "| Idő:" not in text:
        return "", ""
    try:
        rng = text.split("| Idő:", 1)[1].split("|", 1)[0].strip()
        if "-" in rng:
            a, b = rng.split("-", 1)
            return a.strip(), b.strip()
    except Exception:
        pass
    return "", ""


def _clean_predictability_safe(value):
    text = str(value or "")
    if "| Idő:" in text:
        return text.split("| Idő:", 1)[0].strip()
    return text


def build_calendar_view(events_df):
    if events_df is None or events_df.empty:
        return pd.DataFrame(columns=["Nap", "Idő", "Program", "Kiszámíthatóság", "Környezeti tényezők", "Utazás", "Átállás", "Recovery"])

    rows = []
    for _, r in events_df.iterrows():
        raw_pred = str(r.get("kiszamithatosag", r.get("kiszámíthatóság", "")) or "")
        start_key = 9999
        time_text = "nincs megadva"
        if "| Idő:" in raw_pred:
            try:
                rng = raw_pred.split("| Idő:", 1)[1].split("|", 1)[0].strip()
                time_text = rng.replace("-", "–")
                start_raw = rng.split("-", 1)[0].strip()
                h, m = start_raw.split(":")
                start_key = int(h) * 60 + int(m)
            except Exception:
                time_text = "nincs megadva"

        day = r.get("day", r.get("Nap", ""))
        rows.append({
            "Nap": day,
            "Idő": time_text,
            "Program": r.get("program_tipus", r.get("program_típus", "")),
            "Kiszámíthatóság": raw_pred.split("| Idő:", 1)[0].strip(),
            "Környezeti tényezők": r.get("kornyezeti_tenyezok", r.get("környezeti_tényezők", "")),
            "Utazás": f"{r.get('utazas_perc', r.get('utazás_perc', 0))} perc",
            "Átállás": r.get("atallas_szam", r.get("átállás_szám", 0)),
            "Recovery": f"{r.get('recovery_ora', r.get('recovery_óra', 0))} óra",
            "_day": DAYS.index(day) if day in DAYS else 99,
            "_start": start_key,
        })

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values(["_day", "_start"]).drop(columns=["_day", "_start"])


def render_calendar_cards(calendar_df):
    if calendar_df is None or calendar_df.empty:
        st.info("Még nincs naptárba rendezhető program.")
        return
    for day in DAYS:
        day_df = calendar_df[calendar_df["Nap"] == day]
        if day_df.empty:
            continue
        st.markdown(f"#### {day}")
        for _, r in day_df.iterrows():
            st.markdown(
                f"""
                <div class="card">
                    <span class="pill blue">{r['Idő']}</span>
                    <h3>{r['Program']}</h3>
                    <div class="small">
                        <b>Kiszámíthatóság:</b> {r['Kiszámíthatóság']}<br>
                        <b>Környezeti tényezők:</b> {r['Környezeti tényezők']}<br>
                        <b>Utazás:</b> {r['Utazás']} · <b>Átállás:</b> {r['Átállás']} · <b>Recovery:</b> {r['Recovery']}
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )



tab_profile,tab_plan,tab_calendar,tab_checkin,tab_rewards,tab_dash,tab_history,tab_export=st.tabs(["1. Gyermekprofil","2. Heti programtervező","Strukturált naptár","3. Napi check-in","Jutalmazás","4. Stabilitási elemzés","5. Mentett adatok / history","6. Export"])

with tab_profile:
    st.subheader("Gyermekprofil")
    row=children[children["id"]==selected_child_id].iloc[0].to_dict() if selected_child_id else {}
    with st.form("child_form"):
        c1,c2,c3=st.columns(3)
        with c1:
            nickname=st.text_input("Gyermek neve / beceneve", value=row.get("nickname",""))
            age_group=st.selectbox("Életkor",["3–5 év","6–8 év","9–12 év","13+ év"])
            sok_ember_zaj=st.slider("Sok ember / zaj mennyire terhelő?",1,5,int(row.get("sok_ember_zaj",4)), help="1 = alig zavarja, 5 = rövid ideig is erősen leterhelheti a zaj vagy sok ember.")
            screen_sensitivity=st.slider("Képernyőidő mennyire boríthatja a napot?",1,5,int(row.get("screen_sensitivity",3)), help="1 = kevés hatás, 5 = kis plusz képernyőidő is megboríthatja az estét vagy a másnapot.")
        with c2:
            atallas=st.slider("Átállások mennyire nehezek?",1,5,int(row.get("atallas",3)), help="Átállás = elindulás, programváltás, hazaérkezés, egyik helyzetből a másikba lépés.")
            rutinvalidas=st.slider("Váratlan helyzet / rutinváltás mennyire nehéz?",1,5,int(row.get("rutinvalidas",4)), help="Azt jelzi, mennyire nehéz, ha borul a megszokott napirend vagy előre nem jelzett dolog történik.")
            utazas=st.slider("Utazás mennyire terhelő?",1,5,int(row.get("utazas",3)), help="Autó, busz, tömegközlekedés, várakozás, út közbeni ingerek összterhelése.")
            weather_sensitivity=st.slider("Front / időjárás érzékenység",1,5,int(row.get("weather_sensitivity",2)))
        with c3:
            tarsas_helyzet=st.slider("Társas helyzet mennyire merítő?",1,5,int(row.get("tarsas_helyzet",3)))
            recovery_igeny=st.slider("Lecsendesedési idő igénye",1,5,int(row.get("recovery_igeny",4)))
            alvas=st.slider("Rossz alvás mennyire borítja a napot?",1,5,int(row.get("alvas",4)))
            afternoon_sleep_need=st.slider("Délutáni pihenés / alvás igénye",1,5,int(row.get("afternoon_sleep_need",3)))
        segito=st.multiselect("Megnyugvási stratégiák", SETTLE_OPTIONS+["Előre szólás","Fülvédő","Kevesebb beszéd"], default=split_text(row.get("segito_strategiak","")) or ["Csendes szoba","Előre szólás"])
        health=st.multiselect("Állandó vagy gyakori társuló tényezők", HEALTH_OPTIONS, default=split_text(row.get("eu_tarsulo_tenyezok","")) or ["Nincs"])
        save=st.form_submit_button("Gyermekprofil mentése", use_container_width=True)
    if save:
        payload={"nickname":nickname.strip() or "Gyermek","age_group":age_group,"sok_ember_zaj":sok_ember_zaj,"screen_sensitivity":screen_sensitivity,"atallas":atallas,"rutinvalidas":rutinvalidas,"utazas":utazas,"weather_sensitivity":weather_sensitivity,"tarsas_helyzet":tarsas_helyzet,"recovery_igeny":recovery_igeny,"alvas":alvas,"afternoon_sleep_need":afternoon_sleep_need,"segito_strategiak":", ".join(segito),"sajat_megnyugtatas":sajat_megnyugtatas,"eu_tarsulo_tenyezok":", ".join(health)}
        try:
            if selected_child_id: sb.table("children").update(payload).eq("id",selected_child_id).execute()
            else: sb.table("children").insert(payload).execute()
            st.success("Gyermekprofil mentve."); st.rerun()
        except Exception as exc: st.error(f"Mentés sikertelen: {exc}")

if not selected_child_id:
    st.info("Először hozz létre egy gyermekprofilt."); st.stop()
children=load_children(sb); profile=children[children["id"]==selected_child_id].iloc[0].to_dict()

with tab_plan:
    st.subheader("Heti programtervező")
    st.caption("Gyorsított bevitel: óvoda/iskola 5 munkanapra, előző hét másolása, majd finomhangolás.")

    all_events = load_all_events(sb, selected_child_id)
    existing_weeks = sorted(all_events["week_label"].dropna().unique().tolist()) if not all_events.empty and "week_label" in all_events.columns else []
    previous_candidates = [w for w in existing_weeks if w != week_label]

    with st.expander("Előző hét átemelése"):
        if previous_candidates:
            source_week = st.selectbox("Melyik hetet másoljuk?", previous_candidates, index=len(previous_candidates)-1)
            if st.button("Előző hét programjainak átemelése erre a hétre", use_container_width=True):
                try:
                    n = copy_previous_week(sb, selected_child_id, source_week, week_label)
                    st.success(f"{n} program átmásolva. Most már csak finomhangolni kell.")
                    st.rerun()
                except Exception as exc:
                    st.error(f"Másolás sikertelen: {exc}")
        else:
            st.info("Még nincs korábbi hét, amit át lehetne emelni.")

    with st.form("program_form"):
        c1, c2, c3 = st.columns(3)

        with c1:
            day = st.selectbox("Nap", DAYS)
            program_type = st.selectbox("Program típusa", list(PROGRAM_TYPES.keys()))
            add_workdays = st.checkbox("Ha óvoda / iskola: tegye be mind az 5 munkanapra", value=False)
            duration = st.slider("Időtartam (óra)", 0.5, 8.0, 1.0, step=0.5)

        with c2:
            travel = st.slider("Utazás összesen (perc)", 0, 120, 0, step=5)
            transitions = st.slider("Átállások száma", 0, 6, 1)
            predictability = st.selectbox("Kiszámíthatóság", ["Kiszámítható", "Részben kiszámítható", "Váratlan / bizonytalan"])

        with c3:
            envs = st.multiselect("Környezeti tényezők", ENV_OPTIONS, default=["Ismerős hely"])
            recovery_hours = st.slider("Utána tervezett nyugodt blokk (óra)", 0.0, 5.0, 0.5, step=0.5)
            parent_load = st.slider("Szülői szervezési terhelés", 1, 5, 2)

        st.markdown("#### Időpont a strukturált naptárhoz")
        t1, t2, t3 = st.columns(3)
        with t1:
            use_time_range = st.checkbox(
                "Használja az időpontot a naptárban",
                value=True,
                help="Ha kikapcsolod, a program időpont nélkül kerül a naptárba."
            )
        with t2:
            start_time_program = st.time_input("Kezdés", value=time(8, 0), key="program_start_time")
        with t3:
            end_time_program = st.time_input("Befejezés", value=time(9, 0), key="program_end_time")

        submitted = st.form_submit_button("Program mentése felhőbe", use_container_width=True)

    if submitted:
        days_to_insert = WORKDAYS if add_workdays and program_type == "Óvoda / iskola" else [day]
        try:
            time_text = ""
            if use_time_range and start_time_program is not None and end_time_program is not None:
                time_text = f" | Idő: {start_time_program.strftime('%H:%M')}-{end_time_program.strftime('%H:%M')}"

            for d in days_to_insert:
                event_payload = {
                    "child_id": selected_child_id,
                    "week_label": week_label,
                    "day": d,
                    "program_tipus": program_type,
                    "idotartam_ora": duration,
                    "utazas_perc": travel,
                    "atallas_szam": transitions,
                    "kiszamithatosag": predictability + time_text,
                    "kornyezeti_tenyezok": ", ".join(envs),
                    "recovery_ora": recovery_hours,
                    "szuloi_terheles": parent_load,
                }
                event_payload["terhelesi_pont"] = calculate_event_load(event_payload, profile)
                insert_event(sb, event_payload)

            st.success(f"{len(days_to_insert)} program mentve.")
            st.rerun()
        except Exception as exc:
            st.error(f"Mentés sikertelen: {exc}")

    events_df = load_events(sb, selected_child_id, week_label)
    st.markdown("### Heti programok")
    if events_df.empty:
        st.info("Még nincs program ezen a héten.")
    else:
        st.dataframe(normalize_events_df(events_df), use_container_width=True, hide_index=True)
        delete_options = {f"{r['day']} · {r['program_tipus']} · {r['created_at']}": r["id"] for _, r in events_df.iterrows()}
        with st.expander("Program törlése"):
            del_label = st.selectbox("Törlendő program", list(delete_options.keys()))
            if st.button("Kiválasztott program törlése"):
                delete_event(sb, delete_options[del_label])
                st.success("Program törölve.")
                st.rerun()


with tab_calendar:
    st.subheader("Strukturált heti naptár")
    st.caption("A felvitt programokból automatikusan rendezett heti naptár készül. Ha megadsz tól–ig időpontot, időrendben is sorba rendezi.")
    events_for_calendar = load_events(sb, selected_child_id, week_label)
    calendar_df = build_calendar_view(events_for_calendar)

    st.markdown("### Táblázatos naptár")
    st.dataframe(calendar_df, use_container_width=True, hide_index=True)

    st.markdown("### Kártyás naptárnézet")
    render_calendar_cards(calendar_df)


with tab_checkin:
    checkins=load_checkins(sb,selected_child_id,week_label)
    st.subheader("Napi check-in"); st.success(engagement(checkins)); st.info("Új szabály: egy naphoz egy aktív check-in tartozik. Ha ugyanarra a napra újra mentesz, az aznapi korábbi check-in felülíródik.")
    with st.form("checkin_form"):
        c1,c2,c3=st.columns(3)
        with c1:
            fb_day=st.selectbox("Nap",DAYS); morning=st.selectbox("Reggeli ébredés / indulás",list(MORNING_OPTIONS.keys()))
            evening=st.selectbox("Esti állapot",list(FEEDBACK_OPTIONS.keys())); fatigue=st.slider("Napközbeni fáradtság",1,5,3, help="1 = energikus, 5 = nagyon fáradt / alig bírta a napot.")
        with c2:
            sleep=st.slider("Alvás minősége előző éjjel",1,5,3, help="1 = nagyon rossz, 5 = kifejezetten jó. A rossz alvás sokszor előjel."); meal=st.selectbox("Étkezés",list(MEAL_OPTIONS.keys()))
            screen=st.slider("Kütyüidő / képernyőidő (perc)",0,180,30,step=10, help="Becsült képernyőidő. Nem minősítés, csak mintafigyelés."); weather=st.selectbox("Külső tényező",WEATHER_OPTIONS)
        with c3:
            expectation=st.slider("Elvárásokkal töltött idő (óra)",0.0,10.0,4.0,step=0.5, help="Olyan idő, amikor teljesíteni, alkalmazkodni, figyelni kellett: iskola, óvoda, fejlesztés, szabályozott program."); own_rec=st.slider("Saját regenerációs idő / szabad játék (óra)",0.0,6.0,1.0,step=0.5, help="Elvárásmentes, saját tempójú idő. Nem feltétlenül alvás.")
            challenge=st.selectbox("Kihívást jelentő helyzet",CHALLENGE_OPTIONS, help="Ha nem volt nehéz helyzet, hagyd 'Nem volt' értéken. A részletező mezők ilyenkor is maradhatnak 'Nem volt'-on.")
            st.markdown("**Kihívást jelentő helyzet részletei**")
            challenge_time=st.selectbox("Mikor történt?",CHALLENGE_TIME_OPTIONS,index=0,help="A napszak segít felismerni, mikor fogy el gyakrabban a kapacitás.")
            challenge_phase=st.selectbox("Mihez kapcsolódott?",CHALLENGE_PHASE_OPTIONS,index=0,help="Nem mindegy, hogy program előtt, közben, után vagy átálláskor jelent meg.")
            challenge_location=st.selectbox("Hol történt?",CHALLENGE_LOCATION_OPTIONS,index=0)
            challenge_trigger=st.selectbox("Fő kiváltó ok",CHALLENGE_TRIGGER_OPTIONS,index=0,help="Nem kell biztosnak lenni: elég a legvalószínűbb okot jelölni.")
            challenge_duration=st.selectbox("Kb. meddig tartott?",CHALLENGE_DURATION_OPTIONS,index=0)
            settle=st.multiselect("Mi segített?",SETTLE_OPTIONS)
        helped_free=st.text_input("Mi segített még? (opcionális)", placeholder="pl. kedvenc zene, közös kuckózás")
        illness=st.multiselect("Időszakos eü. / állapot",HEALTH_OPTIONS,default=["Nincs"])
        sub=st.form_submit_button("Check-in mentése felhőbe", use_container_width=True)
    if sub:
        save_single_checkin_for_day(sb, {"child_id":selected_child_id,"week_label":week_label,"day":fb_day,"reggeli_allapot":morning,"reggeli_allapot_pont":MORNING_OPTIONS[morning],"esti_allapot":evening,"esti_allapot_pont":FEEDBACK_OPTIONS[evening],"napkozbeni_faradsag":fatigue,"alvas_minosege":sleep,"etkezes":meal,"etkezes_pont":MEAL_OPTIONS[meal],"kutyuidoperc":screen,"kulso_tenyezo":weather,"elvaras_ido":expectation,"sajat_regeneracio_ido":own_rec,"kihivas_helyzet":f"{challenge} | Időszak: {challenge_time} | Kapcsolódás: {challenge_phase} | Helyzet: {challenge_location} | Kiváltó: {challenge_trigger} | Időtartam: {challenge_duration}","megnyugvast_segitette":", ".join(settle),"mi_segitett_szabad_szoveg":helped_free,"idoszakos_eu_allapot":", ".join(illness)})
        st.success("Check-in mentve."); st.rerun()
    st.dataframe(norm_checkins(checkins), use_container_width=True, hide_index=True)
    if not checkins.empty:
        st.markdown("#### Check-in kezelés")
        days_with_checkin = sorted(checkins["day"].dropna().unique().tolist(), key=lambda x: DAYS.index(x) if x in DAYS else 99)
        dc1, dc2 = st.columns(2)
        with dc1:
            del_day = st.selectbox("Melyik nap check-injét töröljük?", days_with_checkin, key="delete_checkin_day_select")
            if st.button("Kiválasztott nap check-injének törlése", use_container_width=True):
                delete_checkins_for_day(sb, selected_child_id, week_label, del_day)
                st.success(f"{del_day} check-in törölve.")
                st.rerun()
        with dc2:
            if st.button("Duplikált check-inek tisztítása ezen a héten", use_container_width=True):
                deleted_count = cleanup_duplicate_checkins_current_week(sb, selected_child_id, week_label)
                st.success(f"Tisztítás kész. Törölt régi duplikátumok: {deleted_count}.")
                st.rerun()


with tab_rewards:
    st.subheader("Jutalmazás és matricagyűjtés")
    st.caption("A cél nem a tökéletes viselkedés, hanem hogy együtt figyeljük és jobban értsük a hetet.")

    rewards_checkins = load_checkins(sb, selected_child_id, week_label)
    points = calculate_reward_points(rewards_checkins)

    st.metric("Összegyűjtött matricák", f"{points} / 7")
    st.success(f"Szint: {reward_level(points)}")

    cols = st.columns(7)
    for i in range(7):
        with cols[i]:
            if i < points:
                st.markdown("⭐")
            else:
                st.markdown("⬜")

    st.markdown("### Választható jutalmak")

    selected_rewards = st.multiselect(
        "Mik motiválják most legjobban?",
        REWARD_OPTIONS
    )

    custom_reward = st.text_input(
        "Saját jutalom ötlet",
        placeholder="pl. új matrica, kis Lego, közös fagyizás"
    )

    if points >= 3:
        st.info("3 matrica után már érdemes egy kisebb jutalmat adni.")
    if points >= 5:
        st.success("5 matrica = erős hét. Jöhet egy nagyobb közös élmény.")
    if points >= 7:
        st.balloons()
        st.success("Teljes hét végigkövetve! Ez hatalmas segítség a mintázatok felismerésében.")

    if selected_rewards:
        st.markdown("### Aktuális jutalomötletek")
        for r in selected_rewards:
            st.markdown(f"- {r}")

    if custom_reward.strip():
        st.markdown("### Saját jutalom")
        st.markdown(f"- {custom_reward}")


with tab_dash:
    events=load_events(sb,selected_child_id,week_label); checkins=load_checkins(sb,selected_child_id,week_label)
    summary=build_day_summary(events,checkins); insights=pd.DataFrame(generate_insights(summary,profile,events,checkins))
    st.subheader("Heti stabilitási elemzés")
    if summary.empty: st.info("Adj hozzá programokat.")
    else:
        risk=clamp(summary["Kockázat"].sum()/330*100); stability=clamp(100-risk+summary["Recovery_órák"].sum()*2); maxday=summary.sort_values("Kockázat",ascending=False).iloc[0]
        c1,c2,c3,c4=st.columns(4); c1.metric("Heti stabilitás",f"{stability}/100",score_color(stability)); c2.metric("Túlterhelési kockázat",f"{risk}/100",score_color(risk,True)); c3.metric("Legnehezebb nap",str(maxday["Nap"]),f"{maxday['Kockázat']:.0f} pont"); c4.metric("Check-in napok", checkins["day"].nunique() if not checkins.empty else 0)
        fig=px.bar(summary,x="Nap",y="Kockázat",color="Kockázat",title="Napi idegrendszeri terhelés / kockázat"); fig.update_layout(template="plotly_dark",xaxis_title="",yaxis_title="Kockázati pont"); st.plotly_chart(fig,use_container_width=True)
        st.dataframe(summary,use_container_width=True,hide_index=True)
        for _,item in insights.iterrows():
            klass="red" if item["Szint"]=="KRITIKUS" else ("yellow" if item["Szint"]=="FIGYELMEZTETÉS" else "green")
            st.markdown(f"""<div class='card'><span class='pill {klass}'>{item['Szint']}</span><h3>{item['Téma']}</h3><b>Mit látunk?</b><br>{item['Mit látunk?']}<br><br><b>Mit érdemes tenni?</b><br>{item['Mit érdemes tenni?']}</div>""",unsafe_allow_html=True)
        st.markdown("### Részletesebb heti következtetések")
        for item in build_deeper_conclusions(summary,profile,events,checkins):
            st.markdown(f"""<div class='card'><span class='pill blue'>KÖVETKEZTETÉS</span><h3>{item['Cím']}</h3><b>Következtetés</b><br>{item['Következtetés']}<br><br><b>Javaslat</b><br>{item['Javaslat']}</div>""",unsafe_allow_html=True)

        st.markdown("### Prognózis / mire figyeljetek előre?")
        prognosis_df = build_prognosis_engine(summary, profile, events, checkins)
        st.dataframe(prognosis_df, use_container_width=True, hide_index=True)

        challenge_df = build_challenge_pattern_table(checkins)
        if not challenge_df.empty:
            st.markdown("### Kihívást jelentő helyzetek mintázatai")
            st.dataframe(challenge_df, use_container_width=True, hide_index=True)

        allc=load_all_checkins(sb,selected_child_id)
        if not allc.empty and allc["week_label"].nunique()>=2:
            trend=allc.groupby("week_label",as_index=False).agg(átlag_esti=("esti_allapot_pont","mean"), átlag_reggeli=("reggeli_allapot_pont","mean"), átlag_kütyüidő=("kutyuidoperc","mean"), checkin_napok=("day","nunique"))
            st.markdown("### Többhetes mintázatok"); st.dataframe(trend,use_container_width=True,hide_index=True)
            fig2=px.line(trend,x="week_label",y=["átlag_esti","átlag_reggeli"],markers=True,title="Állapot trend több héten"); fig2.update_layout(template="plotly_dark"); st.plotly_chart(fig2,use_container_width=True)

with tab_history:
    ae=load_all_events(sb,selected_child_id); ac=load_all_checkins(sb,selected_child_id)
    st.subheader("Mentett adatok / history"); c1,c2,c3=st.columns(3); c1.metric("Mentett programok",len(ae)); c2.metric("Mentett check-inek",len(ac)); c3.metric("Hetek",len(set(list(ae.get("week_label",[]))+list(ac.get("week_label",[])))) if not ae.empty or not ac.empty else 0)
    st.markdown("### Program history"); st.dataframe(norm_events(ae),use_container_width=True,hide_index=True)
    st.markdown("### Check-in history"); st.dataframe(norm_checkins(ac),use_container_width=True,hide_index=True)


def build_full_visual_pdf_report(profile, events_df, day_summary, insights_df, checkins_df, week_label):
    if insights_df is None:
        insights_df = pd.DataFrame()
    """Stabil, teljes heti PDF riport: diagram + minden fő stabilitási blokk."""
    if SimpleDocTemplate is None:
        return None

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=1.1*cm,
        leftMargin=1.1*cm,
        topMargin=1.0*cm,
        bottomMargin=1.0*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("HU_Title", parent=styles["Title"], fontSize=17, leading=21, spaceAfter=8)
    h2_style = ParagraphStyle("HU_H2", parent=styles["Heading2"], fontSize=12, leading=15, textColor=colors.HexColor("#1E3A8A"), spaceBefore=8, spaceAfter=5)
    body_style = ParagraphStyle("HU_Body", parent=styles["Normal"], fontSize=8.4, leading=10.5)
    small_style = ParagraphStyle("HU_Small", parent=styles["Normal"], fontSize=7.5, leading=9)

    def safe_text(value):
        if value is None:
            return ""
        text = str(value)
        # ReportLab alapfont biztonság miatt hosszú magyar ékezetek egyszerűsítése
        return text.replace("ő", "ö").replace("Ő", "Ö").replace("ű", "ü").replace("Ű", "Ü")

    def P(value, style=body_style):
        return Paragraph(safe_text(value), style)

    def add_df_section(story, title, df, max_rows=10):
        story.append(P(title, h2_style))
        if df is None or df.empty:
            story.append(P("Nincs elég adat ehhez a blokkhoz.", small_style))
            story.append(Spacer(1, 0.12*cm))
            return

        for _, row in df.head(max_rows).iterrows():
            vals = []
            for col in df.columns:
                v = row.get(col, "")
                if pd.notna(v) and str(v).strip():
                    vals.append(f"<b>{safe_text(col)}:</b> {safe_text(v)}")
            if vals:
                story.append(P(" | ".join(vals), small_style))
                story.append(Spacer(1, 0.07*cm))

    story = []
    child_name = profile.get("nickname") or profile.get("gyermek_neve") or "Gyermek"

    story.append(P(f"Neurodiverz családi heti riport – {child_name}", title_style))
    story.append(P(f"Hét: {week_label} · Generálva: {datetime.now().strftime('%Y-%m-%d %H:%M')}", small_style))
    story.append(Spacer(1, 0.25*cm))

    if day_summary is None or day_summary.empty:
        story.append(P("Nincs elég adat a heti riporthoz.", body_style))
        doc.build(story)
        return output.getvalue()

    weekly_risk_raw = day_summary["Kockázat"].sum() if "Kockázat" in day_summary.columns else 0
    overload_risk = clamp(weekly_risk_raw / 330 * 100)
    recovery_sum = day_summary["Recovery_órák"].sum() if "Recovery_órák" in day_summary.columns else 0
    stability = clamp(100 - overload_risk + recovery_sum * 2)
    max_day = day_summary.sort_values("Kockázat", ascending=False).iloc[0] if "Kockázat" in day_summary.columns else day_summary.iloc[0]
    checkin_days = checkins_df["day"].nunique() if checkins_df is not None and not checkins_df.empty and "day" in checkins_df.columns else 0

    summary_data = [
        [P("Heti stabilitás", small_style), P("Túlterhelési kockázat", small_style), P("Legnehezebb nap", small_style), P("Check-in napok", small_style)],
        [P(f"{stability}/100", body_style), P(f"{overload_risk}/100", body_style), P(str(max_day.get("Nap", max_day.get("day", ""))), body_style), P(str(checkin_days), body_style)],
    ]
    table = Table(summary_data, colWidths=[4.15*cm, 4.15*cm, 4.15*cm, 4.15*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1E3A8A")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("BACKGROUND", (0,1), (-1,1), colors.HexColor("#EFF6FF")),
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("RIGHTPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.3*cm))

    # Diagram
    try:
        story.append(_make_weekly_bar_drawing(day_summary))
        story.append(Spacer(1, 0.3*cm))
    except Exception as exc:
        story.append(P(f"A heti diagram nem volt elkészíthető: {exc}", small_style))

    # Alap insightok
    add_df_section(story, "Szülőbarát insightok", insights_df, max_rows=8)

    # Részletes következtetések
    try:
        deeper = pd.DataFrame(build_deeper_conclusions(day_summary, profile, events_df, checkins_df))
        add_df_section(story, "Részletesebb heti következtetések", deeper, max_rows=8)
    except Exception as exc:
        story.append(P("Részletesebb heti következtetések", h2_style))
        story.append(P(f"Nem sikerült elkészíteni: {exc}", small_style))

    # Neurodiverz insight engine blokkok
    try:
        pack = build_neurodiverz_insight_engine(day_summary, profile, events_df, checkins_df)
        add_df_section(story, "Heti összkép", pack.get("heti_osszkep", pd.DataFrame()), max_rows=6)
        add_df_section(story, "Több nap fókuszban", pack.get("napi_fokusz", pd.DataFrame()), max_rows=8)
        add_df_section(story, "Mintázatok és összefüggések", pack.get("mintazatok", pd.DataFrame()), max_rows=8)
        add_df_section(story, "Pozitív jelek / ami működni látszik", pack.get("pozitiv_jelek", pd.DataFrame()), max_rows=8)
        add_df_section(story, "Konkrét mikro-javaslatok", pack.get("mikro_javaslatok", pd.DataFrame()), max_rows=8)
    except Exception as exc:
        story.append(P("Neurodiverz insight engine", h2_style))
        story.append(P(f"Nem sikerült elkészíteni: {exc}", small_style))

    # Prognózis
    try:
        prognosis = build_prognosis_engine(day_summary, profile, events_df, checkins_df)
        add_df_section(story, "Prognózis / mire figyeljetek előre?", prognosis, max_rows=8)
    except Exception as exc:
        story.append(P("Prognózis", h2_style))
        story.append(P(f"Nem sikerült elkészíteni: {exc}", small_style))

    # Kihívást jelentő helyzetek
    try:
        challenge_df = build_challenge_pattern_table(checkins_df)
        add_df_section(story, "Kihívást jelentő helyzetek mintázata", challenge_df, max_rows=10)
    except Exception as exc:
        story.append(P("Kihívást jelentő helyzetek", h2_style))
        story.append(P(f"Nem sikerült elkészíteni: {exc}", small_style))

    # Jutalmazás
    try:
        points = calculate_reward_points(checkins_df)
        reward_df = pd.DataFrame([{
            "Matricák": f"{points}/7",
            "Szint": reward_level(points),
            "Üzenet": "A jutalom nem a tökéletes viselkedésért jár, hanem azért, mert együtt figyeltétek a hetet."
        }])
        add_df_section(story, "Jutalmazás / matricagyűjtés", reward_df, max_rows=2)
    except Exception:
        pass

    story.append(P("Fontos megjegyzés", h2_style))
    story.append(P(
        "Ez az eszköz nem diagnosztikai vagy egészségügyi rendszer. "
        "Célja a családi terhelés, rutin, átállások, alvás/étkezés/képernyő és recovery tudatosabb tervezése.",
        small_style
    ))

    doc.build(story)
    return output.getvalue()


with tab_export:
    events=load_events(sb,selected_child_id,week_label); checkins=load_checkins(sb,selected_child_id,week_label); summary=build_day_summary(events,checkins); insights=pd.DataFrame(generate_insights(summary,profile,events,checkins))
    st.subheader("Export")
    if summary.empty: st.info("Nincs exportálható heti elemzés.")
    else:
        st.dataframe(summary,use_container_width=True,hide_index=True); st.dataframe(insights,use_container_width=True,hide_index=True)
        st.download_button("⬇️ Excel riport letöltése", data=export_excel(profile,events,summary,insights,checkins), file_name=f"neurodiverz_csaladi_command_center_v4_5_2_3_6_5_4_{week_label}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        
    try:
        insights_df
    except NameError:
        try:
            insights_df = pd.DataFrame(generate_insights(day_summary, profile, events_df, checkins_df))
        except Exception:
            insights_df = pd.DataFrame()


    pdf_bytes = build_full_visual_pdf_report(profile, events_df, day_summary, insights_df, checkins_df, week_label)
        if pdf_bytes is not None:
            st.download_button("⬇️ Vizuális heti PDF riport letöltése", data=pdf_bytes, file_name=f"neurodiverz_heti_vizualis_riport_v4_5_2_3_6_5_4_{week_label}.pdf", mime="application/pdf", use_container_width=True)
        else:
            st.info("PDF exporthoz a requirements.txt fájlban szerepelnie kell: reportlab és matplotlib")
    st.info("Ez az eszköz nem diagnosztikai vagy egészségügyi rendszer. Célja a családi terhelés és mintázatok tudatosabb követése.")
